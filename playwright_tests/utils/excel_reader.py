from openpyxl import load_workbook
from pathlib import Path
import random


def get_feature_data():
    file_path = Path("test_data/AI_Test_Feature_Data.xlsx")

    workbook = load_workbook(file_path)
    sheet = workbook.active

    max_row = sheet.max_row
    random_row = random.randint(2, max_row)

    feature_summary = sheet.cell(row=random_row, column=1).value
    feature_requirements = sheet.cell(row=random_row, column=2).value

    return feature_summary, feature_requirements